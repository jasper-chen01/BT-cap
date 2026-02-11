"""
Chat agent service for conversational interface
"""
import asyncio
import re
from typing import Optional, List, Dict
from datetime import datetime
from collections import Counter

from backend.models.schemas import ChatMessage, ChatResponse, ChatSession
from backend.services.annotation_service import AnnotationService
from backend.config import settings
from pathlib import Path
from google.cloud import storage
import openpyxl
import os
import json



try:
    import google.generativeai as genai
except Exception:  # pragma: no cover - optional dependency
    genai = None

try:
    import vertexai
    from vertexai.generative_models import GenerativeModel
    from google.oauth2 import service_account
except Exception:  # pragma: no cover - optional dependency
    vertexai = None
    GenerativeModel = None
    service_account = None


class ChatAgent:
    """AI agent for conversational interaction"""
    
    def __init__(self):
        self.annotation_service = AnnotationService()
        self.greetings = [
            "hello", "hi", "hey", "greetings", "good morning", "good afternoon", "good evening"
        ]
        self.help_keywords = ["help", "guide", "tutorial", "instructions"]
        self.annotate_keywords = ["annotate", "analyze", "process", "classify", "label", "predict"]
        self.status_keywords = ["status", "health", "ready", "available"]
        self.gemini_model = None
        self.gemini_enabled = False
        self._init_gemini()

    def _init_gemini(self) -> None:
        """Initialize Gemini client if configured."""
        if settings.VERTEX_PROJECT_ID and vertexai is not None and GenerativeModel is not None:
            credentials = None
            if settings.GOOGLE_APPLICATION_CREDENTIALS and service_account is not None:
                credentials = service_account.Credentials.from_service_account_file(
                    settings.GOOGLE_APPLICATION_CREDENTIALS
                )
            vertexai.init(
                project=settings.VERTEX_PROJECT_ID,
                location=settings.VERTEX_LOCATION,
                credentials=credentials
            )
            self.gemini_model = GenerativeModel(settings.GEMINI_MODEL)
            self.gemini_enabled = True
            return

        if settings.GEMINI_API_KEY and genai is not None:
            genai.configure(api_key=settings.GEMINI_API_KEY)
            self.gemini_model = genai.GenerativeModel(settings.GEMINI_MODEL)
            self.gemini_enabled = True
    
    async def process_message(
        self,
        session: ChatSession,
        user_message: str,
        file_path: Optional[str] = None
    ) -> ChatResponse:
        """
        Process user message and generate appropriate response
        
        Args:
            session: Current chat session
            user_message: User's message text
            file_path: Optional path to uploaded file
        
        Returns:
            ChatResponse with agent's reply
        """
        message_lower = user_message.lower().strip()
        

        
        # Handle greetings
        if any(self._has_keyword(message_lower, greeting) for greeting in self.greetings):
            return self._greeting_response(session.session_id)
        
        # Handle help requests
        if any(self._has_keyword(message_lower, keyword) for keyword in self.help_keywords):
            return self._help_response(session.session_id)
        
        # Handle status checks
        if any(self._has_keyword(message_lower, keyword) for keyword in self.status_keywords):
            return await self._status_response(session.session_id)
        
        # Handle file upload
        if file_path:
            return await self._handle_file_upload(session, file_path, user_message)
        
        # Handle annotation requests
        if any(self._has_keyword(message_lower, keyword) for keyword in self.annotate_keywords):
            if session.uploaded_files:
                # Annotate the most recent file
                return await self._annotate_latest_file(session)
            else:
                return self._request_file_upload(session.session_id)
        
        # Handle explicit annotation requests with file reference
        file_ref_match = re.search(r'file\s*(\d+)', message_lower)
        if file_ref_match and session.uploaded_files:
            file_index = int(file_ref_match.group(1)) - 1
            if 0 <= file_index < len(session.uploaded_files):
                return await self._annotate_specific_file(session, file_index)
        
        # Extract parameters from message
        top_k = self._extract_number(message_lower, r'top[_\s]?k[:\s]?(\d+)', default=10)
        threshold = self._extract_number(message_lower, r'threshold[:\s]?([\d.]+)', default=0.7, is_float=True)
        # 0) Marker weight lookup (runs before Gemini)
        raw = (user_message or "").strip()

        # try to grab a gene-like token (CHRM1, EGFR, OLIG2, etc.)
        gene_match = re.search(r"\b[A-Za-z0-9-]{2,20}\b", raw)

        lookup_query = gene_match.group(0) if gene_match else raw
        

        try:
            lookup = self._lookup_weight_from_xlsx(user_message)  # <-- your existing method name
        except Exception as e:
            print("MARKER LOOKUP ERROR:", repr(e))
            lookup = None
        
        if lookup:
            kind, q, rows = lookup
            
            if rows:
                
                #if kind == "gene":
                   # lines = [f"{q.upper()} weights:"]
                  #  for cell, gene, w in rows[:25]:
                 #       lines.append(f"- {cell}: {w}")
                #else:
                   # lines = [f"{q} marker weights:"]
                  #  for cell, gene, w in rows[:25]:
                 #       lines.append(f"- {gene}: {w}")

                #return ChatResponse(
                  #  message=ChatMessage(role="assistant", content="\n".join(lines)),
                 #   session_id=session.session_id
                #)

                context = {
                    "query_type": kind,
                    "query": q,
                    "results": [{"cell_type": c, "gene": g, "weight": w} for c, g, w in rows[:20]],
                }
                prompt = f"""
                Answer using ONLY these lookup results. If empty, say not found.
                No swearing. Stay on topic. <=3 sentences.

                Lookup results:
                {context}

                User: {user_message}
                """
                resp = await asyncio.to_thread(self.gemini_model.generate_content, prompt)
                content = (resp.text or "").strip() or "Found results but couldn't format them."
                return ChatResponse(message=ChatMessage(role="assistant", content=content), session_id=session.session_id)
                

        # Default response
        if self.gemini_enabled:
            return await self._gemini_response(session, user_message)
        return self._default_response(session.session_id, session.uploaded_files)
    
    async def annotate_file(
        self,
        session: ChatSession,
        file_path: str,
        top_k: int = 10,
        similarity_threshold: float = 0.7
    ) -> ChatResponse:
        """Annotate a file and return chat response"""
        try:
            result = await self.annotation_service.annotate_file(
                file_path,
                top_k=top_k,
                similarity_threshold=similarity_threshold
            )
            
            # Generate summary
            annotations = result.annotations
            annotation_counts = {}
            total_confidence = 0
            
            for cell in annotations:
                ann = cell.predicted_annotation
                annotation_counts[ann] = annotation_counts.get(ann, 0) + 1
                total_confidence += cell.confidence_score
            
            avg_confidence = total_confidence / len(annotations) if annotations else 0
            
            # Create summary text
            summary_lines = [
                f"✅ Successfully annotated {result.total_cells} cells!\n\n",
                f"**Annotation Summary:**\n"
            ]
            
            for ann, count in sorted(annotation_counts.items(), key=lambda x: -x[1]):
                percentage = (count / result.total_cells) * 100
                summary_lines.append(f"- {ann}: {count} cells ({percentage:.1f}%)")
            
            summary_lines.append(f"\n**Average Confidence:** {avg_confidence * 100:.1f}%")
            
            content = "\n".join(summary_lines)
            
            message = ChatMessage(
                role="assistant",
                content=content,
                annotation_results={
                    "total_cells": result.total_cells,
                    "annotation_counts": annotation_counts,
                    "average_confidence": avg_confidence,
                    "annotations": [
                        {
                            "cell_id": cell.cell_id,
                            "predicted_annotation": cell.predicted_annotation,
                            "confidence_score": cell.confidence_score
                        }
                        for cell in annotations[:10]  # Include first 10 for preview
                    ]
                }
            )
            
            return ChatResponse(
                message=message,
                session_id=session.session_id,
                suggestions=[
                    "Download full results",
                    "Visualize annotations",
                    "Upload another file"
                ]
            )
            
        except Exception as e:
            error_message = ChatMessage(
                role="assistant",
                content=f"❌ Error annotating file: {str(e)}\n\nPlease check that:\n- The file is in h5ad format\n- The file contains valid single-cell data\n- Reference embeddings are loaded"
            )
            return ChatResponse(
                message=error_message,
                session_id=session.session_id
            )
    
    def _greeting_response(self, session_id: str) -> ChatResponse:
        """Generate greeting response"""
        content = """Hello! I'm your Brain Tumor Annotation Assistant.\n
I can help you:\n
- Upload and annotate single-cell glioma data\n
- Analyze your data using our reference embeddings\n
- Provide detailed annotation results\n

You can upload a file by dragging it into the chat or typing "upload file".\n How can I help you today?"""
        
        message = ChatMessage(role="assistant", content=content)
        return ChatResponse(
            message=message,
            session_id=session_id,
            suggestions=["Upload a file", "How does this work?", "Check system status"]
        )
    
    def _help_response(self, session_id: str) -> ChatResponse:
        """Generate help response"""
        content = """ How to use the Brain Tumor Annotation Portal:\n

1. Upload Data: Drag and drop a `.h5ad` file or type "upload file"\n
2. Annotate: Say "annotate" or "analyze my data" to process uploaded files\n
3. Customize: Specify parameters like "top k: 20" or "threshold: 0.8"\n
4. Download: Request to download results after annotation\n

Example Commands:\n
- "Upload my glioma data"
- "Annotate with top k 15"
- "Analyze file 1 with threshold 0.75"
- "What's the system status?"

Supported Formats:
- Input: `.h5ad` files (AnnData format)
- Output: CSV with annotations and confidence scores

Need more help? Just ask!"""
        
        message = ChatMessage(role="assistant", content=content)
        return ChatResponse(
            message=message,
            session_id=session_id,
            suggestions=["Upload a file", "Check system status"]
        )
    
    async def _status_response(self, session_id: str) -> ChatResponse:
        """Check system status"""
        status = self.annotation_service.get_status()
        
        if status["reference_loaded"] and status["index_loaded"]:
            content = f"""System Status: Ready

- Reference data: Loaded
- Embeddings index: Ready ({status['index_size']:,} reference cells)
- System: Operational

You can upload and annotate files now!"""
        else:
            ref_status = "Loaded" if status['reference_loaded'] else "Not loaded"
            idx_status = " Ready" if status['index_loaded'] else "Not ready"
            
            content = f"""System Status: Not Ready

- Reference data: {ref_status}
- Embeddings index: {idx_status}

Please ensure reference data is prepared. Run `python scripts/prepare_reference_embeddings.py` if needed."""
        
        message = ChatMessage(role="assistant", content=content)
        return ChatResponse(
            message=message,
            session_id=session_id
        )
    
    async def _handle_file_upload(
        self,
        session: ChatSession,
        file_path: str,
        user_message: str
    ) -> ChatResponse:
        """Handle file upload"""
        filename = session.uploaded_files[-1]["filename"] if session.uploaded_files else "file"
        
        content = f"""📁 **File uploaded successfully!**

File: `{filename}`

You can now:
- Say "annotate" to analyze this file
- Specify parameters: "annotate with top k 15 and threshold 0.8"
- Upload more files for batch processing

Ready to annotate?"""
        
        message = ChatMessage(
            role="assistant",
            content=content,
            file_uploaded=filename
        )
        
        return ChatResponse(
            message=message,
            session_id=session.session_id,
            suggestions=["Annotate this file", "Upload another file"],
            requires_action="annotate"
        )
    
    async def _annotate_latest_file(self, session: ChatSession) -> ChatResponse:
        """Annotate the most recently uploaded file"""
        if not session.uploaded_files:
            return self._request_file_upload(session.session_id)
        
        file_info = session.uploaded_files[-1]
        return await self.annotate_file(session, file_info["path"])
    
    async def _annotate_specific_file(self, session: ChatSession, file_index: int) -> ChatResponse:
        """Annotate a specific file by index"""
        file_info = session.uploaded_files[file_index]
        return await self.annotate_file(session, file_info["path"])
    
    def _request_file_upload(self, session_id: str) -> ChatResponse:
        """Request file upload"""
        content = """No file uploaded yet

Please upload a `.h5ad` file to get started. You can:\n
- Drag and drop a file into the chat\n
- Click the upload button\n
- Or type "upload file" and select a file\n

Once uploaded, I can annotate it for you!"""
        
        message = ChatMessage(role="assistant", content=content)
        return ChatResponse(
            message=message,
            session_id=session_id,
            requires_action="upload_file"
        )
    
    def _default_response(self, session_id: str, uploaded_files: List) -> ChatResponse:
        """Default response when intent is unclear"""
        if uploaded_files:
            content = """I'm not sure what you're asking. Here's what I can help with:

- Annotate files: Say "annotate" or "analyze my data"
- Get help: Ask "how does this work?" or "help"
- Check status: Ask "what's the system status?"

You have uploaded files ready to annotate. Would you like me to analyze them?"""
        else:
            content = """I'm here to help you annotate glioma single-cell data!

Try saying:
- "Upload a file" to get started
- "How does this work?" for instructions
- "What can you do?" to see my capabilities

What would you like to do?"""
        
        message = ChatMessage(role="assistant", content=content)
        suggestions = ["Upload a file", "How does this work?", "Check system status"] if not uploaded_files else ["Annotate files", "Get help"]
        
        return ChatResponse(
            message=message,
            session_id=session_id,
            suggestions=suggestions
        )

    def _build_gemini_prompt(self, session: ChatSession, user_message: str) -> str:
        """Build a prompt with lightweight context for Gemini."""
        recent_messages = session.messages[-6:] if session.messages else []
        history_lines = [
            f"{msg.role}: {msg.content}" for msg in recent_messages if msg.content
        ]
        uploaded_names = [f["filename"] for f in session.uploaded_files] if session.uploaded_files else []
        uploaded_text = ", ".join(uploaded_names) if uploaded_names else "none"

        system_prompt = (
            "You are the Brain Tumor Annotation Assistant for a web app that annotates "
            "glioma single-cell data. Be concise and helpful, and respond naturally to the user."
            "If the user asks about uploading or annotating files, explain the steps clearly."
            "If the request is outside the app scope, say you can only help with this portal."
            "No profanity or swearing. Do not repeat profanity even if the user uses it."
        )

        history_block = "\n".join(history_lines) if history_lines else "No prior messages."
        return (
            f"{system_prompt}\n\n"
            f"Uploaded files: {uploaded_text}\n\n"
            f"Conversation:\n{history_block}\n\n"
            f"User: {user_message}\nAssistant:"
        )

    async def _gemini_response(self, session: ChatSession, user_message: str) -> ChatResponse:
        """Generate a Gemini response for general conversation."""
        prompt = self._build_gemini_prompt(session, user_message)
        try:
            response = await asyncio.to_thread(self.gemini_model.generate_content, prompt)
            content = (response.text or "").strip()
            if not content:
                return self._default_response(session.session_id, session.uploaded_files)
            message = ChatMessage(role="assistant", content=content)
            return ChatResponse(message=message, session_id=session.session_id)
        except Exception as e:
            print("GEMINI ERROR:", repr(e))
            return self._default_response(session.session_id, session.uploaded_files)
    
    def _has_keyword(self, text: str, keyword: str) -> bool:
        """Match keyword as a whole word (or exact phrase if multi-word)."""
        keyword = keyword.strip().lower()
        if not keyword:
            return False
        if " " in keyword:
            return keyword in text
        return re.search(rf"\b{re.escape(keyword)}\b", text) is not None

    
    def _extract_number(self, text: str, pattern: str, default: float, is_float: bool = False) -> float:
        """Extract number from text using regex pattern"""
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return float(match.group(1)) if is_float else int(match.group(1))
        return default
    def _download_gcs_xlsx_if_needed(self) -> Path:
        """
        Downloads the XLSX from GCS once and caches it under data/SuppTable1.xlsx.
        Requires GOOGLE_APPLICATION_CREDENTIALS to be set to your service account JSON.
        """
        gcs_uri = os.getenv("MARKER_WEIGHTS_GCS_URI", "").strip()
        if not gcs_uri.startswith("gs://"):
            return settings.DATA_DIR / "SuppTable1.xlsx"  # fallback path

        local_path = settings.DATA_DIR / "SuppTable1.xlsx"
        if local_path.exists():
            return local_path

        # parse gs://bucket/object
        no_scheme = gcs_uri[len("gs://"):]
        bucket_name, blob_path = no_scheme.split("/", 1)

        client = storage.Client()  # uses GOOGLE_APPLICATION_CREDENTIALS
        blob = client.bucket(bucket_name).blob(blob_path)

        local_path.parent.mkdir(parents=True, exist_ok=True)
        blob.download_to_filename(str(local_path))
        return local_path


    def _lookup_weight_from_xlsx(self, user_text: str):
        """
        If user types a gene -> returns all matching rows (CellType, weight).
        If user types a cell type -> returns all matching rows (Gene, weight).
        """
        q = (user_text or "").strip()
        if not q:
            return None

        xlsx_path = self._download_gcs_xlsx_if_needed()
        
      

        if not xlsx_path.exists():
            return None

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        


        sheet_name = (os.getenv("MARKER_WEIGHTS_SHEET", "") or "").strip()
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        # Expect headers in first row: CellType | MarkerGene | weights
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
  
        


        # Find column indices
        def col_idx(name: str) -> int:
            name_l = name.lower()
            for i, h in enumerate(headers):
                if h.lower() == name_l:
                    return i
            return -1

        i_cell = col_idx("CellType")
        i_gene = col_idx("MarkerGene")
        i_w = col_idx("weights")
        if i_w == -1:
            i_w = col_idx("weight")

        if i_cell == -1 or i_gene == -1 or i_w == -1:
            return ("error", q, [])

        is_gene = (" " not in q) and bool(re.fullmatch(r"[A-Za-z0-9\-]{2,20}", q))

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            cell = (str(r[i_cell]).strip() if r[i_cell] is not None else "")
            gene = (str(r[i_gene]).strip() if r[i_gene] is not None else "")
            w = r[i_w]

            if is_gene:
                if gene.upper() == q.upper():
                    rows.append((cell, gene.upper(), w))
            else:
                if cell.lower() == q.lower():
                    rows.append((cell, gene.upper(), w))
        

        return ("gene" if is_gene else "celltype"), q, rows

    
